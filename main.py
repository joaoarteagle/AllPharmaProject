from openpyxl import load_workbook
# from openpyxl.styles import numbers
from datetime import datetime as dt

data = dt.now()
hora_atual = data.strftime("%Y-%m-%d %H:%M:%S")
wb = load_workbook("backups_allpharma.xlsx")
hora = dt.strptime(hora_atual, "%Y-%m-%d %H:%M:%S")

ws = wb.active




count = 4
for row in ws.iter_rows(min_row=4, min_col=3,max_col=3, max_row=22):
    linhaAnt = ws[f'B{count}'].value
    print(linhaAnt)
   
    for cell in row:
        cell.number_format = "HH:MM:ss"
        if "NULL" in str(linhaAnt):
           cell.value = "SEM BACKUP"
        else: 
            cell.value = hora - linhaAnt
            print(cell.value)
    count+=1
    print(count)

wb.save('teste_backup.xlsx')

    
